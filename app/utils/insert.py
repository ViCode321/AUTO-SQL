import os
import openpyxl
from datetime import datetime

def Excel_convert_insert(file):
    global sheet_principal
    try:
        workbook = openpyxl.load_workbook(file)
        sheet_principal = workbook.active
        return f'Archivo {file.name} procesado con éxito.'
    except Exception as e:
        raise ValueError(f'Error al procesar el archivo: {str(e)}')

# Función para generar código SQL de insert
def insert_sql(base, table):
    global sheet_principal
    sql_script = ""
    if not sheet_principal:
        print("No se pudo", sheet_principal)
        return 1, 'Primero debe subir el archivo excel.', sql_script

    indices = [cell.value for cell in sheet_principal[1]]
    # Encontrar la última fila no vacía
    last_non_empty_row = None
    for row in reversed(list(sheet_principal.iter_rows(min_row=2, max_row=sheet_principal.max_row, values_only=True))):
        if any(row):
            last_non_empty_row = row
            break

    if last_non_empty_row is None:
        return 1, '', sql_script

    # Extraer filas, excluyendo la última fila de tipos de datos
    filas = list(sheet_principal.iter_rows(min_row=2, max_row=sheet_principal.max_row - 1, values_only=True))

    tipos_datos = last_non_empty_row

    sql_script = f"INSERT INTO {base}.{table} ({', '.join(indices)}) VALUES\n"
    valores = []
    error_mostrado = False
    for row in filas:
        # Verificar si la fila está completamente vacía
        if all(value is None for value in row):
            continue
        
        if row == tipos_datos:
            continue

        fila_valores = []

        for value, tipo in zip(row, tipos_datos):
            try:
                if tipo == "int":
                    if isinstance(value, int):
                        fila_valores.append(f"{value}" if value is not None else "NULL")
                    else:
                        fila_valores.append("NULL")
                elif tipo == "float":
                    if isinstance(value, (float, int)):
                        fila_valores.append(f"{value:.2f}" if value is not None else "NULL")
                    else:
                        fila_valores.append("NULL")
                elif tipo == "string":
                    fila_valores.append(f"'{str(value)}'" if value is not None else "NULL")
                elif tipo == "date":
                    if isinstance(value, (datetime, datetime.date)):
                        fecha_formateada = value.strftime('%Y-%m-%d %H:%M:%S')
                        fila_valores.append(f"'{fecha_formateada}'" if value is not None else "NULL")
                    else:
                        fila_valores.append("NULL")
                elif tipo == "null":
                    fila_valores.append("NULL")
                else:
                    fila_valores.append(f"'{str(value)}'" if value is not None else "NULL")
            except (ValueError, TypeError) as e:
                fila_valores.append("NULL")
                print(f"Error al convertir el valor '{value}' del tipo '{tipo}': {e}")
        
        valores.append(f"({', '.join(fila_valores)})")

    if valores:
        sql_script += ",\n".join(valores) + ";"
    else:
        sql_script += "NULL;"

    return 0, "", sql_script


sheet_principal = None